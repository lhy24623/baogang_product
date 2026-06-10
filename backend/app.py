"""
Flask后端API服务
RESTful风格接口，按操作类型传参
"""
from flask import Flask, jsonify, request, send_from_directory
from backend.db import db
from backend.config import FLASK_HOST, FLASK_PORT, FLASK_DEBUG

app = Flask(__name__, static_folder='../front', static_url_path='')


@app.route('/')
def index():
    """返回前端页面"""
    return send_from_directory('../front', 'index.html')


@app.route('/product')
def product_page():
    """返回产品档案页面"""
    return send_from_directory('../front', 'product.html')


# ==================== 查询接口 ====================

@app.route('/api/select', methods=['POST'])
def select():
    """
    查询数据
    请求体: {"table": "表名", "columns": ["col1", "col2"], "where": "条件", "limit": 100}
    """
    data = request.get_json() # 从HTTP请求中获取JSON格式的数据
    if not data or 'table' not in data: # 检查数据是否存在，或者数据中是否缺少必需的'table'键
        return jsonify({'success': False, 'error': '请提供表名'}), 400 # 若校验不通过，返回400错误及提示信息

    table = data['table'] # 从请求数据中提取目标表名
    columns = ', '.join(data.get('columns', ['*'])) # 获取查询列，若未指定则默认为'*'，并用逗号拼接成字符串
    where = f"WHERE {data['where']}" if data.get('where') else '' # 若存在查询条件，则拼接WHERE子句，否则为空字符串
    limit = f"LIMIT {data['limit']}" if data.get('limit') else '' # 若存在限制数量，则拼接LIMIT子句，否则为空字符串

    sql = f"SELECT {columns} FROM {table} {where} {limit}".strip() # 拼接完整的SQL查询语句，并去除首尾多余空格
    success, result = db.execute(sql) # 执行SQL语句，并接收执行状态和结果或错误信息
    if success: # 判断数据库执行是否成功
        return jsonify({'success': True, 'data': result, 'sql': sql}) # 成功则返回200状态码及查询数据和SQL语句
    return jsonify({'success': False, 'error': result, 'sql': sql}), 500 # 失败则返回500服务器内部错误及错误信息和SQL语句


# ==================== 插入接口 ====================

@app.route('/api/insert', methods=['POST'])
def insert():
    """
    插入数据
    请求体: {"table": "表名", "data": {"col1": "val1", "col2": "val2"}}
    """
    data = request.get_json()
    if not data or 'table' not in data or 'data' not in data:
        return jsonify({'success': False, 'error': '请提供表名和数据'}), 400

    table = data['table']
    columns = ', '.join(data['data'].keys())
    values = ', '.join([f"'{v}'" if v is not None else 'NULL' for v in data['data'].values()])

    sql = f"INSERT INTO {table} ({columns}) VALUES ({values})"
    success, result = db.execute(sql)
    if success:
        return jsonify({'success': True, 'affected': result, 'sql': sql})
    return jsonify({'success': False, 'error': result, 'sql': sql}), 500


# ==================== 更新接口 ====================

@app.route('/api/update', methods=['POST'])
def update():
    """
    更新数据
    请求体: {"table": "表名", "data": {"col1": "new_val"}, "where": "id=1"}
    """
    data = request.get_json()
    if not data or 'table' not in data or 'data' not in data or 'where' not in data:
        return jsonify({'success': False, 'error': '请提供表名、数据和条件'}), 400

    table = data['table']
    set_clause = ', '.join([f"{k} = '{v}'" if v is not None else f"{k} = NULL" for k, v in data['data'].items()])

    sql = f"UPDATE {table} SET {set_clause} WHERE {data['where']}"
    success, result = db.execute(sql)
    if success:
        return jsonify({'success': True, 'affected': result, 'sql': sql})
    return jsonify({'success': False, 'error': result, 'sql': sql}), 500


# ==================== 删除接口 ====================

@app.route('/api/delete', methods=['POST'])
def delete():
    """
    删除数据
    请求体: {"table": "表名", "where": "id=1"}
    """
    data = request.get_json()
    if not data or 'table' not in data or 'where' not in data:
        return jsonify({'success': False, 'error': '请提供表名和条件'}), 400

    sql = f"DELETE FROM {data['table']} WHERE {data['where']}"
    success, result = db.execute(sql)
    if success:
        return jsonify({'success': True, 'affected': result, 'sql': sql})
    return jsonify({'success': False, 'error': result, 'sql': sql}), 500


# ==================== 原始SQL接口 ====================

@app.route('/api/sql', methods=['POST'])
def raw_sql():
    """
    执行原始SQL（用于SHOW DATABASES等特殊语句）
    请求体: {"sql": "SHOW DATABASES"}
    """
    data = request.get_json()
    if not data or 'sql' not in data:
        return jsonify({'success': False, 'error': '请提供SQL语句'}), 400

    sql = data['sql'].strip()
    if not sql:
        return jsonify({'success': False, 'error': 'SQL语句不能为空'}), 400

    success, result = db.execute(sql)
    if success:
        if isinstance(result, list):
            return jsonify({'success': True, 'data': result, 'sql': sql})
        return jsonify({'success': True, 'affected': result, 'sql': sql})
    return jsonify({'success': False, 'error': result, 'sql': sql}), 500


# ==================== 产品基本信息档案 ====================

# 产品表名
PRODUCT_TABLE = 'tmp_analysis.mid_product_base_info_d_f'

# 筛选字段配置: {前端参数名: (数据库列名, 显示名称)}
FILTER_FIELDS = {
    'prodclass': ('prodclass', '产品大类编码'),
    'prodclasschin': ('prodclasschin', '产品大类名称'),
    'prodtype': ('prodtype', '产品形态编码'),
    'prodtypechin': ('prodtypechin', '产品形态名称'),
}


@app.route('/api/product/filters', methods=['GET'])
def product_filters():
    """获取产品表各筛选字段的去重选项列表，一次查询所有字段"""
    cols = ', '.join(col for col, _ in FILTER_FIELDS.values())
    sql = f"SELECT {cols} FROM {PRODUCT_TABLE}"
    success, result = db.execute(sql)
    options = {}
    if success and result:
        for key, (col, _) in FILTER_FIELDS.items():
            values = sorted(set(row[col] for row in result if row[col] is not None))
            options[key] = values
    else:
        for key in FILTER_FIELDS:
            options[key] = []
    return jsonify({'success': True, 'options': options})


@app.route('/api/product/list', methods=['POST'])
def product_list():
    """
    查询产品基本信息列表
    请求体: {"prodclass": "xxx", "prodclasschin": "xxx", "prodtype": "xxx", "prodtypechin": "xxx", "page": 1, "pageSize": 50}
    """
    data = request.get_json() or {}
    page = data.get('page', 1)
    page_size = data.get('pageSize', 50)
    offset = (page - 1) * page_size

    # 构建WHERE条件
    conditions = []
    for key, (col, _) in FILTER_FIELDS.items():
        val = data.get(key)
        if val:
            conditions.append(f"{col} = '{val}'")
    where = f"WHERE {' AND '.join(conditions)}" if conditions else ''

    # 查询总数
    count_sql = f"SELECT COUNT(*) AS total FROM {PRODUCT_TABLE} {where}"
    success, count_result = db.execute(count_sql)
    if not success:
        return jsonify({'success': False, 'error': count_result}), 500
    total = count_result[0]['total'] if count_result else 0

    # 查询数据
    data_sql = f"SELECT * FROM {PRODUCT_TABLE} {where} LIMIT {page_size} OFFSET {offset}"
    success, result = db.execute(data_sql)
    if not success:
        return jsonify({'success': False, 'error': result}), 500

    return jsonify({
        'success': True,
        'data': result,
        'total': total,
        'page': page,
        'pageSize': page_size
    })


# ==================== 自定义新产品 ====================

CUSTOM_PRODUCT_TABLE = 'tmp_analysis.mid_custom_new_product_info'

# 自定义产品各字段配置: (数据库列名, 显示名称, 是否从原产品表取选项)
CUSTOM_PRODUCT_FIELDS = [
    ('psrno', '产品规范编码', False),
    ('prodclass', '产品大类编码', True),
    ('prodclasschin', '产品大类中文名称', True),
    ('prodtype', '产品形态编码', True),
    ('prodtypechin', '产品形态中文名称', True),
    ('mscno', '制造标准编码', True),
    ('apnno', '用途码', True),
    ('apndesc', '用途中文描述', True),
    ('standname', '标准全名', False),
    ('custom_mark', '自定义产品标记', False),
]

# 需要从原产品表取选项的字段（在原产品表中对应的列名）
CUSTOM_PRODUCT_FILTER_MAP = {
    'prodclass': 'prodclass',
    'prodclasschin': 'prodclasschin',
    'prodtype': 'prodtype',
    'prodtypechin': 'prodtypechin',
    'mscno': 'mscno',
    'apnno': 'apnno',
    'apndesc': 'apndesc',
}


@app.route('/custom-product')
def custom_product_page():
    """返回自定义产品页面"""
    return send_from_directory('../front', 'custom_product.html')


@app.route('/api/custom-product/options', methods=['GET'])
def custom_product_options():
    """获取自定义产品表单的下拉选项，一次查询所有字段"""
    cols = ', '.join(CUSTOM_PRODUCT_FILTER_MAP.values())
    sql = f"SELECT {cols} FROM {PRODUCT_TABLE}"
    success, result = db.execute(sql)
    options = {}
    if success and result:
        for field, col in CUSTOM_PRODUCT_FILTER_MAP.items():
            values = sorted(set(row[col] for row in result if row[col] is not None and row[col] != ''))
            options[field] = values
    else:
        for field in CUSTOM_PRODUCT_FILTER_MAP:
            options[field] = []
    # custom_mark 固定选项
    options['custom_mark'] = ['虚拟产品', '新增实际产品']
    return jsonify({'success': True, 'options': options})


@app.route('/api/custom-product/add', methods=['POST'])
def custom_product_add():
    """
    新增自定义产品
    请求体: {"psrno": "xxx", "prodclass": "xxx", ...}
    """
    data = request.get_json()
    if not data or not data.get('psrno'):
        return jsonify({'success': False, 'error': '产品规范编码(psrno)为必填项'}), 400

    # 自动填充字段
    data['create_time'] = 'NOW()'
    data['source_from'] = '人工定义'

    # 构建INSERT语句，空值用NULL
    columns = []
    values = []
    for col, val in data.items():
        columns.append(col)
        if val is None or val == '':
            values.append('NULL')
        elif val == 'NOW()':
            values.append(val)
        else:
            values.append(f"'{val}'")

    sql = f"INSERT INTO {CUSTOM_PRODUCT_TABLE} ({', '.join(columns)}) VALUES ({', '.join(values)})"
    success, result = db.execute(sql)
    if success:
        return jsonify({'success': True, 'affected': result, 'sql': sql})
    return jsonify({'success': False, 'error': result, 'sql': sql}), 500


# 自定义产品可同步到产品基本信息底表的字段映射
CUSTOM_PRODUCT_SYNC_MAP = {
    'prodclass': 'prodclass',
    'prodclasschin': 'prodclasschin',
    'prodtype': 'prodtype',
    'prodtypechin': 'prodtypechin',
    'mscno': 'mscno',
    'apnno': 'apnno',
    'apndesc': 'apndesc',
}


@app.route('/api/custom-product/sync', methods=['POST'])
def custom_product_sync():
    """
    将自定义新产品表中特定psrno的数据同步到产品基本信息底表
    请求体: {"psrno": "xxx"}
    逻辑: 先查自定义表记录，再检查底表是否已存在该psrno，存在则更新，不存在则插入
    """
    data = request.get_json()
    if not data or not data.get('psrno'):
        return jsonify({'success': False, 'error': '请提供产品规范编码(psrno)'}), 400

    psrno = data['psrno']

    # 1. 从自定义表查询该psrno的记录
    select_sql = f"SELECT * FROM {CUSTOM_PRODUCT_TABLE} WHERE psrno = '{psrno}'"
    success, result = db.execute(select_sql)
    if not success:
        return jsonify({'success': False, 'error': f'查询自定义表失败: {result}'}), 500
    if not result or len(result) == 0:
        return jsonify({'success': False, 'error': f'自定义表中未找到psrno={psrno}的记录'}), 404

    custom_record = result[0]

    # 2. 检查底表中是否已存在该psrno
    check_sql = f"SELECT psrno FROM {PRODUCT_TABLE} WHERE psrno = '{psrno}'"
    success, check_result = db.execute(check_sql)
    if not success:
        return jsonify({'success': False, 'error': f'查询底表失败: {check_result}'}), 500

    # 构建需要同步的字段值（从自定义记录中提取）
    sync_data = {}
    sync_data['psrno'] = psrno
    for custom_col, base_col in CUSTOM_PRODUCT_SYNC_MAP.items():
        if custom_col in custom_record and custom_record[custom_col] is not None:
            sync_data[base_col] = custom_record[custom_col]

    if check_result and len(check_result) > 0:
        # 底表已存在该psrno，执行更新
        set_parts = []
        for col, val in sync_data.items():
            if col == 'psrno':
                continue  # psrno是条件字段，不更新
            if val is None or val == '':
                set_parts.append(f"{col} = NULL")
            else:
                set_parts.append(f"{col} = '{val}'")
        if not set_parts:
            return jsonify({'success': False, 'error': '没有可同步的字段'}), 400
        sql = f"UPDATE {PRODUCT_TABLE} SET {', '.join(set_parts)} WHERE psrno = '{psrno}'"
        action = '更新'
    else:
        # 底表不存在该psrno，执行插入
        columns = []
        values = []
        for col, val in sync_data.items():
            columns.append(col)
            if val is None or val == '':
                values.append('NULL')
            else:
                values.append(f"'{val}'")
        sql = f"INSERT INTO {PRODUCT_TABLE} ({', '.join(columns)}) VALUES ({', '.join(values)})"
        action = '插入'

    success, result = db.execute(sql)
    if success:
        return jsonify({'success': True, 'action': action, 'affected': result, 'sql': sql, 'psrno': psrno})
    return jsonify({'success': False, 'error': result, 'sql': sql}), 500


@app.route('/api/custom-product/delete', methods=['POST'])
def custom_product_delete():
    """
    删除自定义新产品记录
    请求体: {"psrno": "xxx"}
    """
    data = request.get_json()
    if not data or not data.get('psrno'):
        return jsonify({'success': False, 'error': '请提供产品规范编码(psrno)'}), 400

    psrno = data['psrno']
    sql = f"DELETE FROM {CUSTOM_PRODUCT_TABLE} WHERE psrno = '{psrno}'"
    success, result = db.execute(sql)
    if success:
        if result == 0:
            return jsonify({'success': False, 'error': f'未找到psrno={psrno}的记录'}), 404
        return jsonify({'success': True, 'affected': result, 'psrno': psrno})
    return jsonify({'success': False, 'error': result}), 500


@app.route('/api/custom-product/list', methods=['POST'])
def custom_product_list():
    """查询自定义产品列表"""
    data = request.get_json() or {}
    page = data.get('page', 1)
    page_size = data.get('pageSize', 50)
    offset = (page - 1) * page_size

    conditions = []
    for key in ['psrno', 'prodclass', 'prodclasschin', 'prodtype', 'prodtypechin', 'custom_mark']:
        val = data.get(key)
        if val:
            conditions.append(f"{key} = '{val}'")
    where = f"WHERE {' AND '.join(conditions)}" if conditions else ''

    count_sql = f"SELECT COUNT(*) AS total FROM {CUSTOM_PRODUCT_TABLE} {where}"
    success, count_result = db.execute(count_sql)
    if not success:
        return jsonify({'success': False, 'error': count_result}), 500
    total = count_result[0]['total'] if count_result else 0

    data_sql = f"SELECT * FROM {CUSTOM_PRODUCT_TABLE} {where} ORDER BY create_time DESC LIMIT {page_size} OFFSET {offset}"
    success, result = db.execute(data_sql)
    if not success:
        return jsonify({'success': False, 'error': result}), 500

    return jsonify({'success': True, 'data': result, 'total': total, 'page': page, 'pageSize': page_size})


# ==================== 产销数据统计 ====================

# 结算表名
PENDING_TABLE = 'bg_isale.pending_settlement_ss10'
SETTLED_TABLE = 'bg_isale.settlement_ss15'

# 产销统计维度配置: (数据库列名, 显示名称, 筛选类型)
# 筛选类型: select=下拉选择, text=文本输入, date_range=日期范围, time_period=时间聚合, fixed=固定选项
PS_DIMENSIONS = [
    ('stat_period', '时间维度', 'time_period'),
    ('factory', '制造厂', 'select'),
    ('oemVenCode', '委托加工厂', 'select'),
    ('subFactoryNo', '钢联二级厂矿', 'select'),
    ('factoryLine', '生产线', 'select'),
    ('contChannel', '销售渠道', 'select'),
    ('prodCode', '产品代码', 'select'),
    ('prodClass', '产品大类', 'select'),
    ('tradeNo', '钢筋钢种_材质', 'select'),
    ('psrNo', '产品规范代码', 'select'),
    ('specMark', '产品规格描述', 'text'),
    ('apnno', '用途码', 'select'),
    ('settle_status', '结算状态', 'fixed'),
]

# 需要从数据库获取下拉选项的维度列名
PS_SELECT_DIMENSIONS = [col for col, _, ftype in PS_DIMENSIONS if ftype == 'select']

# 时间聚合尺度配置: {尺度key: (SELECT/GROUP BY表达式, 显示名称)}
TIME_GRANULARITY_MAP = {
    'year': ("DATE_FORMAT(shipDate, '%%Y')", '统计年份'),
    'quarter': ("CONCAT(DATE_FORMAT(shipDate, '%%Y'), '-Q', QUARTER(shipDate))", '统计季度'),
    'month': ("DATE_FORMAT(shipDate, '%%Y-%%m')", '统计月份'),
    'week': ("CONCAT(DATE_FORMAT(shipDate, '%%Y'), '-W', WEEK(shipDate))", '统计周'),
    'day': ("DATE_FORMAT(shipDate, '%%Y-%%m-%%d')", '统计日期'),
}

# 统计指标配置: (指标key, SQL聚合表达式, 显示名称)
PS_METRICS = [
    ('order_count', 'COUNT(DISTINCT orderNo)', '订单总数'),
    ('order_item_count', 'COUNT(DISTINCT CONCAT(orderNo, orderItem))', '订单项次总数'),
    ('total_deli_qty', 'SUM(deliQty)', '总出货数量'),
    ('total_deli_wet', 'SUM(deliWet)', '总出货重量'),
    ('total_pack_wet', 'SUM(packWet)', '总包装重量'),
    ('total_discount_qty', 'SUM(discountQty)', '总折让数量'),
    ('total_discount_wet', 'SUM(discountWet)', '总折让重量'),
    ('total_tai_amt', 'SUM(taiAmt)', '台币货款总金额'),
    ('total_frn_amt', 'SUM(frnAmt)', '外币货款总金额'),
    ('total_tax_amt', 'SUM(taxAmt)', '总税额'),
    ('total_adv_charge', 'SUM(advCharge)', '总订金金额'),
    ('total_interest_amt', 'SUM(interestAmt)', '总付款附价'),
    ('total_discount_amt', 'SUM(discountAmt)', '总现金折扣未税'),
    ('total_trans_cost', 'SUM(transCost)', '总运费'),
    ('total_pack_fee', 'SUM(packFee)', '总捆绑费'),
    ('total_port_cost', 'SUM(portCost)', '总站港杂费'),
    ('total_insurance', 'SUM(insurance)', '总SS保费'),
    ('total_agent_trans_cost', 'SUM(agentTransCost)', '代收代付运费总额'),
    ('total_car_price_amount', 'SUM(carPriceAmount)', '自备车租金总额'),
    ('total_shipping_costs', 'SUM(shippingcosts)', '预估海运费总额'),
    ('total_commission', 'SUM(commission)', '预估佣金总额'),
]

# 虚拟产品字段 → 产销统计筛选维度映射
VIRTUAL_PRODUCT_FILTER_MAP = {
    'prodclass': 'prodClass',
    'prodtype': 'prodCode',
    'apnno': 'apnno',
}

# UNION ALL 子查询中待结算/已结算共用的列名（不含settle_status）
PS_INNER_COLUMNS = """shipDate, factory, oemVenCode, subFactoryNo, factoryLine, contChannel,
    prodCode, prodClass, tradeNo, psrNo, specMark, apnno,
    orderNo, orderItem, deliQty, deliWet, packWet, discountQty, discountWet,
    taiAmt, frnAmt, taxAmt, advCharge, interestAmt, discountAmt,
    transCost, packFee, portCost, insurance, agentTransCost,
    carPriceAmount, shippingcosts, commission"""


@app.route('/production-sales')
def production_sales_page():
    """返回产销统计页面"""
    return send_from_directory('../front', 'production_sales.html')


@app.route('/api/production-sales/options', methods=['GET'])
def production_sales_options():
    """获取产销统计各筛选字段的去重选项列表"""
    options = {}
    # 对每个下拉维度，从两张表分别查DISTINCT再合并
    for col in PS_SELECT_DIMENSIONS:
        sql = f"""
            SELECT DISTINCT {col} FROM {PENDING_TABLE} WHERE {col} IS NOT NULL AND {col} != ''
            UNION
            SELECT DISTINCT {col} FROM {SETTLED_TABLE} WHERE {col} IS NOT NULL AND {col} != ''
        """
        success, result = db.execute(sql)
        if success and result:
            options[col] = sorted(set(row[col] for row in result if row[col] is not None and row[col] != ''))
        else:
            options[col] = []
    # settle_status 固定选项
    options['settle_status'] = ['待结算', '已结算']
    # 时间聚合尺度选项
    granularity_options = [{'key': k, 'label': v[1]} for k, v in TIME_GRANULARITY_MAP.items()]
    return jsonify({'success': True, 'options': options, 'dimensions': PS_DIMENSIONS, 'metrics': [(k, l) for k, _, l in PS_METRICS], 'timeGranularity': granularity_options})


@app.route('/api/production-sales/query', methods=['POST'])
def production_sales_query():
    """
    执行产销统计动态查询
    请求体: {
        "filters": {"factory": "xxx", "prodClass": "yyy", ...},
        "groupBy": ["stat_period", "factory", ...],
        "timeGranularity": "month",
        "dateStart": "2026-01-01",
        "dateEnd": "2026-06-30",
        "page": 1,
        "pageSize": 50
    }
    """
    data = request.get_json() or {}
    filters = data.get('filters', {})
    group_by = data.get('groupBy', [])
    time_granularity = data.get('timeGranularity', 'month')
    date_start = data.get('dateStart', '')
    date_end = data.get('dateEnd', '')
    page = data.get('page', 1)
    page_size = data.get('pageSize', 50)
    offset = (page - 1) * page_size

    if not date_start or not date_end:
        return jsonify({'success': False, 'error': '请选择日期范围'}), 400
    if not group_by:
        return jsonify({'success': False, 'error': '请至少选择一个分组维度'}), 400

    # 验证时间聚合尺度
    if time_granularity not in TIME_GRANULARITY_MAP:
        time_granularity = 'month'
    time_expr, time_label = TIME_GRANULARITY_MAP[time_granularity]

    # 构建UNION ALL子查询
    inner_cols_pending = PS_INNER_COLUMNS + ", '待结算' AS settle_status"
    inner_cols_settled = PS_INNER_COLUMNS + ", '已结算' AS settle_status"
    base_query = f"SELECT {inner_cols_pending} FROM {PENDING_TABLE} UNION ALL SELECT {inner_cols_settled} FROM {SETTLED_TABLE}"

    # 构建SELECT子句: 分组维度 + 聚合指标
    select_parts = []
    for dim in group_by:
        if dim == 'stat_period':
            select_parts.append(f"{time_expr} AS `{time_label}`")
        else:
            label = next((l for c, l, _ in PS_DIMENSIONS if c == dim), dim)
            select_parts.append(f"`{dim}` AS `{label}`")
    for _, expr, label in PS_METRICS:
        select_parts.append(f"{expr} AS `{label}`")

    # 构建WHERE子句
    where_parts = [f"shipDate BETWEEN '{date_start}' AND '{date_end}'", "shipDate IS NOT NULL"]
    for dim, val in filters.items():
        if not val:
            continue
        if dim == 'settle_status':
            where_parts.append(f"settle_status = '{val}'")
        elif dim == 'specMark':
            where_parts.append(f"specMark LIKE '%{val}%'")
        else:
            where_parts.append(f"`{dim}` = '{val}'")

    # 构建GROUP BY子句
    group_by_parts = []
    for dim in group_by:
        if dim == 'stat_period':
            group_by_parts.append(time_expr)
        else:
            group_by_parts.append(f"`{dim}`")

    # 构建ORDER BY子句
    order_parts = []
    if 'stat_period' in group_by:
        order_parts.append(f'`{time_label}` ASC')
    if 'settle_status' in group_by:
        order_parts.append('settle_status DESC')

    where_clause = ' AND '.join(where_parts)
    sql_data = f"SELECT {', '.join(select_parts)} FROM ({base_query}) t_total WHERE {where_clause} GROUP BY {', '.join(group_by_parts)}"
    if order_parts:
        sql_data += f" ORDER BY {', '.join(order_parts)}"
    sql_data += f" LIMIT {page_size} OFFSET {offset}"

    # 查询总数
    sql_count = f"SELECT COUNT(*) AS total FROM ({base_query}) t_total WHERE {where_clause} GROUP BY {', '.join(group_by_parts)}"
    sql_count_wrap = f"SELECT COUNT(*) AS total FROM ({sql_count}) sub"

    success, count_result = db.execute(sql_count_wrap)
    if not success:
        return jsonify({'success': False, 'error': count_result, 'sql': sql_count_wrap}), 500
    total = count_result[0]['total'] if count_result else 0

    success, result = db.execute(sql_data)
    if not success:
        return jsonify({'success': False, 'error': result, 'sql': sql_data}), 500

    # 计算维度列和指标列的显示名（用于前端排列列顺序）
    dimension_labels = []
    for dim in group_by:
        if dim == 'stat_period':
            dimension_labels.append(time_label)
        else:
            label = next((l for c, l, _ in PS_DIMENSIONS if c == dim), dim)
            dimension_labels.append(label)
    metric_labels = [label for _, _, label in PS_METRICS]

    return jsonify({
        'success': True,
        'data': result,
        'total': total,
        'page': page,
        'pageSize': page_size,
        'sql': sql_data,
        'dimensionLabels': dimension_labels,
        'metricLabels': metric_labels
    })


@app.route('/api/production-sales/validate', methods=['POST'])
def production_sales_validate():
    """
    数据校验：比对系统统计结果与人工报数
    请求体: {
        "systemData": [{"统计月份": "2026-01", "制造厂": "xxx", "订单总数": 100, ...}],
        "manualData": [{"统计月份": "2026-01", "制造厂": "xxx", "订单总数": 100, ...}],
        "groupBy": ["stat_period", "factory", ...],
        "timeGranularity": "month",
        "threshold": 5
    }
    三层校验：
    1. 总条数一致性
    2. 维度组合一致性（漏报/多报条目）
    3. 字段级校验（以人工为基准：遗漏字段、多报字段、数值差异）
    """
    data = request.get_json() or {}
    system_data = data.get('systemData', [])
    manual_data = data.get('manualData', [])
    group_by = data.get('groupBy', [])
    time_granularity = data.get('timeGranularity', 'month')
    threshold = float(data.get('threshold', 5))

    if not system_data:
        return jsonify({'success': False, 'error': '请先执行查询获取系统统计数据'}), 400
    if not manual_data:
        return jsonify({'success': False, 'error': '请提供人工报数数据'}), 400
    if not group_by:
        return jsonify({'success': False, 'error': '请提供分组维度'}), 400

    # 确定维度列名（结果中的显示名）
    if time_granularity not in TIME_GRANULARITY_MAP:
        time_granularity = 'month'
    _, time_label = TIME_GRANULARITY_MAP[time_granularity]

    dimension_labels = []
    for dim in group_by:
        if dim == 'stat_period':
            dimension_labels.append(time_label)
        else:
            label = next((l for c, l, _ in PS_DIMENSIONS if c == dim), dim)
            dimension_labels.append(label)

    # ---- 第一层：总条数校验 ----
    total_system = len(system_data)
    total_manual = len(manual_data)
    row_count_match = total_system == total_manual

    # ---- 第二层：维度组合校验 ----
    # 取人工报数与系统查询共有的维度列作为匹配依据
    system_keys_set = set(system_data[0].keys()) if system_data else set()
    manual_keys_set = set(manual_data[0].keys()) if manual_data else set()
    match_dims = [dim for dim in dimension_labels if dim in manual_keys_set]

    if not match_dims:
        return jsonify({
            'success': False,
            'error': f'人工报数与查询结果无共有维度列，无法匹配。系统维度: {", ".join(dimension_labels)}，人工报数含: {", ".join(sorted(manual_keys_set))}',
        }), 400

    # 构建维度组合索引（1对1，维度组合应唯一）
    system_dim_map = {}
    for row in system_data:
        key = '|||'.join(str(row.get(dim, '') or '').strip() for dim in match_dims)
        system_dim_map[key] = row

    manual_dim_map = {}
    for row in manual_data:
        key = '|||'.join(str(row.get(dim, '') or '').strip() for dim in match_dims)
        manual_dim_map[key] = row

    system_dim_keys = set(system_dim_map.keys())
    manual_dim_keys = set(manual_dim_map.keys())
    matched_keys = system_dim_keys & manual_dim_keys
    extra_keys = manual_dim_keys - system_dim_keys   # 人工多报的维度组合
    miss_keys = system_dim_keys - manual_dim_keys     # 人工漏报的维度组合

    # ---- 第三层：字段级校验 ----
    # 确定系统列和人工列
    system_columns = sorted(system_keys_set)
    manual_columns = sorted(manual_keys_set)
    common_columns = sorted(system_keys_set & manual_keys_set)
    missing_in_manual = sorted(system_keys_set - manual_keys_set)   # 人工遗漏的字段
    extra_in_manual = sorted(manual_keys_set - system_keys_set)     # 人工多报的字段

    # 判断字段是否为数值型（采样判断）
    def is_numeric_column(rows, col):
        count = 0
        for r in rows[:20]:
            v = r.get(col)
            if v is not None and v != '' and not isinstance(v, bool):
                try:
                    float(v)
                    count += 1
                except (ValueError, TypeError):
                    pass
        return count > min(len(rows), 20) * 0.5

    # 对匹配的维度组合，逐条做字段级校验
    details = []

    # 人工报数条目（以人工为基准）
    for m_row in manual_data:
        dim_key = '|||'.join(str(m_row.get(dim, '') or '').strip() for dim in match_dims)
        s_row = system_dim_map.get(dim_key)

        if s_row is None:
            # 多报条目
            field_checks = {}
            for col in manual_columns:
                field_checks[col] = {'status': 'extra_in_manual', 'manualVal': m_row.get(col)}
            details.append({
                'rowStatus': 'extra',
                'manualRow': m_row,
                'systemRow': None,
                'fieldChecks': field_checks
            })
        else:
            # 匹配条目，逐字段校验
            field_checks = {}
            for col in common_columns:
                m_val = m_row.get(col)
                s_val = s_row.get(col)
                # 数值型字段：比对差异
                if is_numeric_column(system_data, col) and is_numeric_column(manual_data, col):
                    try:
                        m_num = float(m_val) if m_val is not None and m_val != '' else 0.0
                    except (ValueError, TypeError):
                        m_num = 0.0
                    try:
                        s_num = float(s_val) if s_val is not None and s_val != '' else 0.0
                    except (ValueError, TypeError):
                        s_num = 0.0
                    diff_val = m_num - s_num
                    if s_num != 0:
                        diff_rate = (diff_val / abs(s_num)) * 100
                    else:
                        diff_rate = 999.0 if m_num != 0 else 0.0
                    is_diff = abs(diff_rate) > threshold
                    field_checks[col] = {
                        'status': 'diff' if is_diff else 'match',
                        'manualVal': m_num,
                        'systemVal': s_num,
                        'diff': round(diff_val, 2),
                        'diffRate': round(diff_rate, 2)
                    }
                else:
                    # 非数值型字段：比对是否一致
                    m_str = str(m_val or '').strip()
                    s_str = str(s_val or '').strip()
                    field_checks[col] = {
                        'status': 'match' if m_str == s_str else 'diff',
                        'manualVal': m_val,
                        'systemVal': s_val
                    }
            # 人工遗漏的字段
            for col in missing_in_manual:
                field_checks[col] = {'status': 'missing_in_manual', 'systemVal': s_row.get(col)}
            # 人工多报的字段
            for col in extra_in_manual:
                field_checks[col] = {'status': 'extra_in_manual', 'manualVal': m_row.get(col)}

            details.append({
                'rowStatus': 'matched',
                'manualRow': m_row,
                'systemRow': s_row,
                'fieldChecks': field_checks
            })

    # 漏报条目（系统有但人工没有）
    for s_row in system_data:
        dim_key = '|||'.join(str(s_row.get(dim, '') or '').strip() for dim in match_dims)
        if dim_key in miss_keys:
            field_checks = {}
            for col in system_columns:
                field_checks[col] = {'status': 'missing_in_manual', 'systemVal': s_row.get(col)}
            details.append({
                'rowStatus': 'miss',
                'manualRow': None,
                'systemRow': s_row,
                'fieldChecks': field_checks
            })

    return jsonify({
        'success': True,
        'summary': {
            'totalSystem': total_system,
            'totalManual': total_manual,
            'rowCountMatch': row_count_match,
            'matchedCombos': len(matched_keys),
            'extraCombos': len(extra_keys),
            'missCombos': len(miss_keys),
            'commonColumns': common_columns,
            'missingInManual': missing_in_manual,
            'extraInManual': extra_in_manual
        },
        'details': details,
        'dimensionColumns': dimension_labels,
        'matchDimensions': match_dims
    })


@app.route('/api/custom-product/virtual', methods=['GET'])
def custom_product_virtual():
    """
    根据psrno获取虚拟产品配置，返回映射后的筛选条件和分组维度
    请求参数: ?psrno=xxx
    """
    psrno = request.args.get('psrno', '').strip()
    if not psrno:
        return jsonify({'success': False, 'error': '请提供psrno参数'}), 400

    sql = f"SELECT * FROM {CUSTOM_PRODUCT_TABLE} WHERE psrno = '{psrno}' AND custom_mark = '虚拟产品'"
    success, result = db.execute(sql)
    if not success:
        return jsonify({'success': False, 'error': result}), 500
    if not result or len(result) == 0:
        return jsonify({'success': False, 'error': f'未找到psrno={psrno}的虚拟产品配置'}), 404

    record = result[0]

    # 根据映射关系，提取筛选条件和分组维度
    filters = {}
    group_by = []
    for custom_field, ps_dim in VIRTUAL_PRODUCT_FILTER_MAP.items():
        val = record.get(custom_field)
        if val is not None and val != '':
            filters[ps_dim] = val
            group_by.append(ps_dim)

    return jsonify({
        'success': True,
        'psrno': psrno,
        'filters': filters,
        'groupBy': group_by,
        'rawRecord': record
    })


# ==================== 统一导出Excel ====================

@app.route('/api/export/excel', methods=['POST'])
def export_excel():
    """
    统一Excel导出接口
    请求体: {
        "fileName": "导出文件名(不含扩展名)",
        "sheets": [
            {
                "sheetName": "Sheet1",
                "columns": ["列1", "列2", ...],
                "data": [[val1, val2, ...], ...]
            },
            ...
        ]
    }
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO

    body = request.get_json() or {}
    file_name = body.get('fileName', '导出数据')
    sheets = body.get('sheets', [])

    if not sheets:
        return jsonify({'success': False, 'error': '请提供导出数据'}), 400

    wb = Workbook()
    # 删除默认创建的Sheet
    default_sheet = wb.active

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    cell_alignment = Alignment(horizontal='left', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for idx, sheet_data in enumerate(sheets):
        sheet_name = sheet_data.get('sheetName', f'Sheet{idx + 1}')
        columns = sheet_data.get('columns', [])
        data = sheet_data.get('data', [])

        if idx == 0:
            ws = default_sheet
            ws.title = sheet_name[:31]  # Excel sheet名最长31字符
        else:
            ws = wb.create_sheet(title=sheet_name[:31])

        # 写表头
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 写数据
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = cell_alignment
                cell.border = thin_border

        # 自动调整列宽
        for col_idx, col_name in enumerate(columns, 1):
            max_len = len(str(col_name)) * 2  # 中文字符宽度约为英文2倍
            for row_data in data:
                if col_idx - 1 < len(row_data):
                    val = row_data[col_idx - 1]
                    if val is not None:
                        val_len = len(str(val))
                        # 简单估算：中文字符按2计算
                        char_len = sum(2 if ord(c) > 127 else 1 for c in str(val))
                        max_len = max(max_len, char_len)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 50)

    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    from flask import send_file
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{file_name}.xlsx'
    )


# ==================== 健康检查 ====================

@app.route('/api/health', methods=['GET'])
def health():
    """健康检查"""
    if db.connect():
        db.close()
        return jsonify({'status': 'healthy', 'database': 'connected'})
    return jsonify({'status': 'unhealthy', 'database': 'disconnected'}), 500


def run_server():
    """启动服务"""
    print(f"启动服务: http://{FLASK_HOST}:{FLASK_PORT}")
    app.run(host=FLASK_HOST, port=FLASK_PORT, debug=FLASK_DEBUG)


if __name__ == '__main__':
    run_server()